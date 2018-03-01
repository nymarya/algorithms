from openpyxl import load_workbook
import numpy as np
from sklearn.decomposition import PCA

#Load data from workbook
wb = load_workbook('DL03_Teste01_Dados.xlsx')
sheet = wb[u'Planilha1']

#Recover data
datasheet = []
for row in sheet.iter_rows(min_row=3, min_col=2):    
    newRow = list() 
    for cell in row:
       # print type(cell.value)
        newRow.append(cell.value)
    datasheet.append(tuple(newRow))
    
#Transpose matrix so we can calculate both covariance and correlation matrix
data = np.reshape( np.array(datasheet), (9, 683) )
covMatrix = np.cov(data)
corMatrix = np.corrcoef(data)

#get principal components
pca = PCA()
pca.fit(np.array(datasheet))
components = pca.components_

#Getting eigenvectors and eigenvalues
centered_matrix = data - data.mean(axis=1)[:, np.newaxis]
cov = np.dot(centered_matrix, centered_matrix.T)
eigenvalues, eigenvectors = np.linalg.eig(cov)

#Save data in file
datafile = open("analysis.txt", "w")
datafile.write("Original\n")
datafile.write("Covariance matrix: \n")
for i in range(covMatrix.shape[0]):
    datafile.write(str(covMatrix[i]))
    datafile.write('\n')
datafile.close()

datafile = open("analysis.txt", "a")
datafile.write("\n covariance matrix2: \n")
for i in range(cov.shape[0]):
    datafile.write(str(cov[i]))
    datafile.write('\n')

datafile.write("\nCorrelation matrix: \n")
for i in range(corMatrix.shape[0]):
    datafile.write(str(corMatrix[i]))
    datafile.write('\n')

datafile.write("\nPCA: \n")
for i in range(components.shape[0]):
    datafile.write(str(components[i]))
    datafile.write('\n')

#print eigenvectors
datafile.write("\nEigenvectors: \n")
for i in range(eigenvectors.shape[0]):
    datafile.write(str(eigenvectors[i]))
    datafile.write('\n')

#print eigenvalues
datafile.write("\nEigenvalues: \n")
for i in range(eigenvalues.shape[0]):
    datafile.write(str(eigenvalues[i]))
    datafile.write('\n')
datafile.close()

print "Done ;)\nBye"
