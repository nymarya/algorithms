from openpyxl import load_workbook
import numpy as np 

def read ( ):
    #Load data from workbook
    wb = load_workbook('DL03_Teste01_Dados.xlsx')
    sheet = wb['Planilha1']

    #Recover data
    datasheet = []
    for row in sheet.iter_rows():  
        newRow = list()
        for cell in row:
            newRow.append(cell.value)
        datasheet.append(tuple(newRow))

    #training portion
    training_set_size_portion = .8

    # Dados randomicos
    randomized_data  = np.random.permutation(datasheet)

    # Definir tamanho de conjuntos
    total_records = len(randomized_data)
    training_set_size = int(total_records * training_set_size_portion)
    test_set_size = total_records - training_set_size

    # Pegar conjunto de treino
    training_set = []
    resp_training_set = []
    for i in range(0, training_set_size):
        tuple_size = len(datasheet[i])
        feautures = datasheet[i][0:tuple_size-1] 	
        training_set.append(feautures)
        resp_training_set.append(datasheet[i][tuple_size-1])

    # pegar conjunto de teste
    test_set = []
    resp_test_set = []
    for i in range(training_set_size+1, total_records ):
        tuple_size = len(datasheet[i])
        feautures = datasheet[i][0:tuple_size-1] 	
        test_set.append(feautures)
        resp_test_set.append(datasheet[i][tuple_size-1])

    return [training_set, resp_training_set, test_set, resp_test_set]
