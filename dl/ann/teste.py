from openpyxl import load_workbook
import numpy as np
from sklearn.decomposition import PCA
import cov
import math

#Load data from workbook
wb1 = load_workbook('ConjTreino.xlsx')
sheet1 = wb1[u'Sheet1']

wb2 = load_workbook('ConjTeste.xlsx')
sheet2 = wb2[u'Sheet1']

#Recover data
datasheet = []
for row in sheet1.iter_rows(min_row=3, min_col=2, max_col=10):    
    newRow = list() 
    for cell in row:
        newRow.append(cell.value)
    datasheet.append(tuple(newRow))
    
#Transpose matrix so we can calculate both covariance and correlation matrix
dados = np.array(datasheet)

from sklearn.preprocessing import StandardScaler
train_data_std = StandardScaler().fit_transform(dados)

#get covariance matrix
#egenvalues and eigenvectors
cov_mat = np.cov(train_data_std.T)
eig_vals, eig_vecs = np.linalg.eig(cov_mat)

print('Eigenvectors \n%s' %eig_vecs)
print('\nEigenvalues \n%s' %eig_vals)

for ev in eig_vecs:
    np.testing.assert_array_almost_equal(1.0, np.linalg.norm(ev))

# Make a list of (eigenvalue, eigenvector) tuples
eig_pairs = [(np.abs(eig_vals[i]), eig_vecs[:,i]) for i in range(len(eig_vals))]

# Sort the (eigenvalue, eigenvector) tuples from high to low
eig_pairs.sort()
eig_pairs.reverse()

# Visually confirm that the list is correctly sorted by decreasing eigenvalues
print('Eigenvalues in descending order:')
for i in eig_pairs:
    print(i[0])

tot = sum(eig_vals)
cum_var_exp = 0.0
for i in sorted(eig_vals, reverse=True):
    var_exp = (i / tot)*100 
    cum_var_exp += var_exp
    print('var %f cumvar %f' %(var_exp, cum_var_exp))

#reduce the dimension

matrix_w = np.hstack((eig_pairs[0][1].reshape(9,1), 
                      eig_pairs[1][1].reshape(9,1),
                      eig_pairs[2][1].reshape(9,1), 
                      eig_pairs[3][1].reshape(9,1),
                      eig_pairs[4][1].reshape(9,1), 
                      eig_pairs[5][1].reshape(9,1),
                      eig_pairs[6][1].reshape(9,1),
                      ))
                      
print('Matrix W:\n', matrix_w.shape)

training_data = train_data_std.dot(matrix_w)

#Recover data
datasheet = []
for row in sheet2.iter_rows(min_row=3, min_col=2, max_col=10):    
    newRow = list() 
    for cell in row:
        newRow.append(cell.value)
    datasheet.append(tuple(newRow))

test_data = np.array(datasheet)    
test_data_std = StandardScaler().fit_transform(test_data)
test_data = test_data_std.dot(matrix_w)


## get target
#treino
datasheet = []
for row in sheet1.iter_rows(min_row=3, min_col=12, max_col=12):     
    for cell in row:
        datasheet.append(cell.value)
    
target_train = np.array([ 1 if i==2 else 0 for i in datasheet])

datasheet = []
for row in sheet2.iter_rows(min_row=3, min_col=12, max_col=12):     
    for cell in row:
        datasheet.append(cell.value)
    
target_test = np.array([ 1 if i==2 else 0 for i in datasheet])


#from sklearn.model_selection import train_test_split
#X_train, X_test, y_train, y_test = train_test_split(X, y)

#training

from sklearn.neural_network import MLPClassifier

#create 3 layers with 30 neurons each
mlp = MLPClassifier(hidden_layer_sizes=(300, 500), batch_size=100)
mlp.fit(training_data,target_train)

#prediction
predictions = mlp.predict(test_data)

#prin confusion matrix
from sklearn.metrics import classification_report,confusion_matrix
print(confusion_matrix(target_test,predictions))

print(classification_report(target_test,predictions))


print ("Done ;)\nBye")
